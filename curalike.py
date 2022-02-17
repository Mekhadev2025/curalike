fil1="products.xlsx"                                         
fil2="meta.xlsx"                                               
fil3="match.xlsx"                              
r1=508                                                      
r2=66                                                        

 
from openpyxl import Workbook
wrkbk1=Workbook()
s1=wrkbk1.active
s1["A1"]="ID"                                                   
s1["B1"]="Key"                                                  
s1["C1"]="Value"                                              






from openpyxl import load_workbook
wrkbk2=load_workbook(filename=fil1)
s2=wrkbk2.active

 
pdict={}
for r in range(2,r1):
    dc=s2.cell(row=r, column=3).value 
                      
    if dc!=None:                                              
        id=s2.cell(row=r, column=1).value                       
        pdict[id]=dc                                          


 
wrkbk3=load_workbook(filename=fil2)
s3=wrkbk3.active

 
dict2={}
for r in range(2,r2):
    pvalue=s3.cell(row=r, column=2).value                      
    pkey=s3.cell(row=r, column=1).value                          
    dict2[pvalue]=pkey                                          


 
num=2
for c in pdict:                                                 
    id=c                                                         
    p=pdict[c]                                                                                  
    for d in dict2:                                             
        key= dict[2d]                                           
        value=d                                                 
        if value.lower() in p.lower():                          
            s1.cell(row=num, column=3).value=value               
            s1.cell(row=num, column=2).value=key                 
            s1.cell(row=num, column=1).value=id                  
            num+=1



wrkbk1.save(fil3)                                                    

print("Program executed")
    